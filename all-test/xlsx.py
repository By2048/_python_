import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl import load_Workbook

import os

table=xlrd.open_workbook('book.xlsx')

# 获取所有的表名
print(table.sheet_names())

table_sheet=table.sheet_by_index(0)

# 递归打印出每行的信息
for row in range(table_sheet.nrows):
    print(table_sheet.row_values(row))

# rowx 行    colx 列
data00=table_sheet.cell(0,0).value
data01=table_sheet.cell(rowx=1,colx=2).value
print(data00)
print(data01)

write_table=xlwt.Workbook('book_write.xlsx')
sheet1=write_table.add_sheet('book')
write_table.save('book_write.xls')
