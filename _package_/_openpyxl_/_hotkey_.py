import string
import os

from openpyxl import load_workbook

file = r"E:\Sync\Ahk\Doc\Hotkey.xlsx"
from openpyxl import load_workbook

wb = load_workbook(filename=file)
# print(wb.sheetnames)


sheet = wb["|1|"]

print(sheet["A1"].value)

for y in range(1, 10):
    for x in list(string.ascii_uppercase):
        print(f"{x}{y}", end=" ")
    print()

# cell_range = sheet['A1':'D11']

# for item in cell_range:
#     print(sheet[item])
#
# print(cell_range)


# print(sheet['D18'].value)
