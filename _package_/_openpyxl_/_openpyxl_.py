import logging

from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('_openpyxl_.xlsx')


def test_read():
    def test_1():
        for name in wb.sheetnames:
            for row in wb[name].rows:
                logging.info([col.value for col in row])

    def test_2():
        for name in wb.sheetnames:
            for col in wb[name].columns:
                logging.info([row.value for row in col])

    test_1()
    test_2()


def test_write():
    # TODO write excel
    pass


if __name__ == '__main__':
    test_read()
