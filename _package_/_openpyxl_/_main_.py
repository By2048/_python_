#### python 操控 excel

from openpyxl import Workbook
from openpyxl import load_workbook

# 加载文件
worl_excel = load_workbook('flash.xlsx')

# 获取所有表名
sheets = worl_excel.get_sheet_names()

# 选择表
work_sheet = worl_excel.get_sheet_by_name("Sheet1")

# 获取活动表
work_sheet = worl_excel.active

# 获取行列数据
row_len = work_sheet.rows
col_len = work_sheet.columns

sheet_rows = []
sheet_cols = []


# 获取所有数据 行
def get_sheet_rows():
    for name in sheet_names:
        work_sheet = worl_excel.get_sheet_by_name(name)
        tmp = []
        for row in ws.rows:
            tmp.append([col.value for col in row])
        sheet_rows.append(tmp)


# 获取所有数据 列
def get_sheet_cols():
    for name in sheet_names:
        work_sheet = worl_excel.get_sheet_by_name(name)
        tmp = []
        for col in ws.columns:
            tmp.append([row.value for row in col])
        sheet_cols.append(tmp)


sheet.title = 'myname'
print(sheet.title)

# 根据单元格数据设置/获取值
sheet['A1'] = 'Id'
print(sheet['A1'].value)

# 根据行列设置/获取值
sheet.cell(row=1, column=2).value = 'Name'
print(sheet.cell(row=1, column=2).value)

# 直接插入一行
work_sheet.append([1, 2, 3])

# 操作完成后需要保存文件
worl_excel.save(filename='create.xlsx')
