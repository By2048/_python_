from openpyxl import load_workbook
from openpyxl import Workbook

wb=load_workbook('e:\\desktop\\test.xlsx')
# ws = wb.active


# 所有的行列
sheet_rows=[]
sheet_cols=[]

sheet_names=wb.get_sheet_names()

def get_sheet_rows():
    for name in sheet_names:
        ws =wb.get_sheet_by_name(name)
        tmp=[]
        for row in ws.rows:
            tmp.append([col.value for col in row])
        sheet_rows.append(tmp)

def get_sheet_cols():
    for name in sheet_names:
        ws =wb.get_sheet_by_name(name)
        tmp=[]
        for col in ws.columns:
            tmp.append([row.value for row in col])
        sheet_cols.append(tmp)

get_sheet_rows()
get_sheet_cols()


def set_1_col():
    ws = wb.get_sheet_by_name(sheet_names[3])
    for row in range(2,len(sheet_cols[0][0][1:])+2):
        num = sheet_cols[0][0][row-1]
        ws.cell(row=row,column=1).value=num
        print(ws.cell(row=row, column=1).value)

def set_2_5_col():
    sheet1_infos = sheet_cols[0][3][1:]
    sheet2_col_infos = sheet_cols[1][3][1:]
    sheet2_row_infos=sheet_rows[1]

    ws4 = wb.get_sheet_by_name(sheet_names[3])

    tmps = []
    for info1 in sheet1_infos:
        info2_col=2
        for info2 in sheet2_col_infos:
            tmp=[]
            if info1==info2:
                print(sheet2_row_infos[info2_col][0:5])
                tmp=sheet2_row_infos[info2_col][0:5]
                tmps.append(tmp)
            info2_col += 1
    print(tmps)

    for row in range(2, len(sheet_cols[0][0][1:]) + 2):
        for col in range(2,7):
            ws4.cell(row=row,column=col).value=tmps[row-2][col-2]

    for row in range(2, len(sheet_cols[0][0][1:]) + 2):
        for col in range(2, 7):
            print(ws4.cell(row=row, column=col).value,end='\t')
        print('\n')


set_1_col();
set_2_5_col()


wb.save('e:\\desktop\\test.xlsx')
